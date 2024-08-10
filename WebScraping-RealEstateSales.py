from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl
import os
from datetime import datetime

# Definindo as constantes
DOMAIN = 'https://endereço_do_site'
URL = f'{DOMAIN}/alugar/imovel/aguas-claras-brasilia-df-brasil'
OUTPUT_FILE = 'dados_imoveis.xlsx'

options = Options()
options.add_argument("--disable-javascript")
options.add_argument("--disable-gpu")
options.add_argument("--disable-images")
options.add_argument("--headless")
options.add_argument("--max-old-space-size=4096")

# Criando o driver do Chrome
navegador = webdriver.Chrome(options=options)
navegador.maximize_window()

navegador.get(URL)
contadorPagina = 1
quantidade_de_imoveis = None

try:
    QuantDeImoveis = navegador.find_elements(By.CSS_SELECTOR, 'p.CozyTypography.xih2fc.Bm9DLt.EqjlRj span')
    for elemento in QuantDeImoveis:
        try:
            quantidade_de_imoveis = elemento.text
            quantidade_de_imoveis = int(quantidade_de_imoveis)
        except:
            quantidade_de_imoveis = elemento.text.replace('.', '')
            quantidade_de_imoveis = int(quantidade_de_imoveis)
        print(f'{elemento.text} Imóveis nessa localidade')
except Exception as e:
    print(f"Erro ao obter a quantidade de imóveis: {e}")

while True:
    try:
        ver_mais_button = navegador.execute_script(
            "return document.querySelector('button.Cozy__Button-Component.bvqY2e.wVqmS5.LLYiJC[aria-label=\"Ver mais\"]');")
        if ver_mais_button:
            navegador.execute_script("arguments[0].click();", ver_mais_button)
            WebDriverWait(navegador, 15).until(EC.staleness_of(ver_mais_button))
            contadorPagina += 1
            numeroDePaginas = (quantidade_de_imoveis / 12)
            print(f'{contadorPagina} páginas lidas de {numeroDePaginas:.2f}')
        else:
            break
    except Exception as e:
        print(f'Erro: {e}')
        break

try:
    tipo = navegador.find_elements(By.CSS_SELECTOR, 'div.Zkjoo- h2')
    valores = navegador.find_elements(By.CSS_SELECTOR, 'div.Cozy__CardTitle-Title.hFUhPy h3')
    valorAluguel = navegador.find_elements(By.CSS_SELECTOR, 'div.Cozy__CardTitle-Subtitle.JyjznE h3')
    endereco = navegador.find_elements(By.CSS_SELECTOR, 'div.Cozy__CardContent-Container.XBxUCJ h3')
    enderecoPreciso = navegador.find_elements(By.CSS_SELECTOR, 'div.Cozy__CardContent-Container.XBxUCJ h2')

    if os.path.exists(OUTPUT_FILE):
        planilha = openpyxl.load_workbook(OUTPUT_FILE)
        aba = planilha.active
        row = aba.max_row + 1
    else:
        planilha = openpyxl.Workbook()
        aba = planilha.active
        aba['A1'] = 'Tipo'
        aba['B1'] = 'Valor Total'
        aba['C1'] = 'Valor Aluguel'
        aba['D1'] = 'Informações'
        aba['E1'] = 'Endereço'
        aba['F1'] = 'Data e Hora da consulta'
        row = 2
except Exception as e:
    print(f'Erro, continuando de onde parou {e}')

for i in range(len(valores)):
    try:
        print(f"Dados salvos {i + 1}")
        aba.cell(row=row, column=1, value=tipo[i].text)
        aba.cell(row=row, column=2, value=valores[i].text)
        aba.cell(row=row, column=3, value=valorAluguel[i].text)
        aba.cell(row=row, column=4, value=endereco[i].text)
        aba.cell(row=row, column=5, value=enderecoPreciso[i].text)
        aba.cell(row=row, column=6, value=datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        row += 1
    except Exception as e:
        print(f'Salvando...O que não foi perdido {e}')

planilha.save(OUTPUT_FILE)
print(f"DADOS SALVOS NO ARQUIVO: {OUTPUT_FILE}")
navegador.quit()
