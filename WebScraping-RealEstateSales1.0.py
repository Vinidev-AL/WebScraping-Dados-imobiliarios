from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
import openpyxl


driver = webdriver.Chrome()

# Base URL for scraping
base_url = "https://www.dfimoveis.com.br/venda/df/brasilia/apartamento?pagina={}"

# Create a new workbook
workbook = openpyxl.Workbook()
sheet = workbook.active

# Write the headers
sheet['A1'] = 'Página'
sheet['B1'] = 'Elemento'
sheet['C1'] = 'Valor'
sheet['D1'] = 'Metros'
sheet['E1'] = 'Valor Metro Quadrado'
sheet['F1'] = 'Quartos'
sheet['G1'] = 'Suites'
sheet['H1'] = 'Vagas'
sheet['I1'] = 'Endereço'

# Loop through different pages
row = 2  # Start from row 2 (row 1 is for headers)
for page in range(1, 3):  # Iterate over pages 1 to 168
    url = base_url.format(page)
    driver.get(url)

    for i in range(1, 41):  # Iterate over elements on each page (1 to 40)
        try:
            try:
                valor_xpath = f'//*[@id="resultadoDaBuscaDeImoveis"]/a[{i}]/div[2]/div[2]/div[1]/h4[1]/span'
                valor = driver.find_element(By.XPATH, valor_xpath).text
            except:
                valor = 'Sem valor'
                
            try:
                metros_xpath = f'//*[@id="resultadoDaBuscaDeImoveis"]/a[{i}]/div[2]/div[2]/ul/li[1]/span'
                metros = driver.find_element(By.XPATH, metros_xpath).text
            except:
                metros = 'Sem metros'
                
            try:
                valorMetro_xpath = f'//*[@id="resultadoDaBuscaDeImoveis"]/a[{i}]/div[2]/div[2]/div[1]/h4[2]/span'
                valor_Metro_Quadrado = driver.find_element(By.XPATH, valorMetro_xpath).text
            except:
                valor_Metro_Quadrado = 'Sem valor'
                
            try:
                quartos_xpath = f'//*[@id="resultadoDaBuscaDeImoveis"]/a[{i}]/div[2]/div[2]/ul/li[2]/span'
                quartos = driver.find_element(By.XPATH, quartos_xpath).text
            except:
                quartos = 'Sem quartos'
                
            try:
                suites_xpath = f'//*[@id="resultadoDaBuscaDeImoveis"]/a[{i}]/div[2]/div[2]/ul/li[3]/span'
                suites = driver.find_element(By.XPATH, suites_xpath).text
            except:
                suites = 'Sem suites'
                
            try:
                vagas_xpath = f'//*[@id="resultadoDaBuscaDeImoveis"]/a[{i}]/div[2]/div[2]/ul/li[4]/span'
                vagas = driver.find_element(By.XPATH, vagas_xpath).text
            except:
                vagas = 'Sem vagas'
                
            try:
                endereco_xpath = f'//*[@id="resultadoDaBuscaDeImoveis"]/a[{i}]/div[2]/div[1]/div/h2'
                endereco = driver.find_element(By.XPATH, endereco_xpath).text
            except:
                endereco = 'Sem endereço'

            # Write data to the Excel sheet
            sheet.cell(row=row, column=1, value=page)
            sheet.cell(row=row, column=2, value=i)
            sheet.cell(row=row, column=3, value=valor)
            sheet.cell(row=row, column=4, value=metros)
            sheet.cell(row=row, column=5, value=valor_Metro_Quadrado)
            sheet.cell(row=row, column=6, value=quartos)
            sheet.cell(row=row, column=7, value=suites)
            sheet.cell(row=row, column=8, value=vagas)
            sheet.cell(row=row, column=9, value=endereco)

            row += 1  # Move to the next row
            print(f"Successfully scraped data from page {page}, element {i}")

        except Exception as e:
            print(f"Error scraping page {page}, element {i}: {str(e)}")
            continue  # Continue to the next element after encountering an exception

# Save the workbook to a file
output_file = "sales_scraped_data.xlsx"
workbook.save(output_file)
print(f"All scraped data saved to {output_file}")

# Close the browser
driver.quit()